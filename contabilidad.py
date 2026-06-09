import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
import database
import theme


class Contabilidad:

    def __init__(self, master):
        self.frame = theme.frame(master)
        self.frame.columnconfigure(0, weight=1)
        self.frame.rowconfigure(2, weight=1)

        # Título
        title_bar = theme.frame(self.frame, bg=theme.BG_CARD)
        title_bar.grid(row=0, column=0, sticky="ew")
        theme.label(title_bar, "📊  Contabilidad", style="title", bg=theme.BG_CARD).pack(side="left", padx=20, pady=14)

        # ── Controles ────────────────────────────────
        ctrl = theme.card(self.frame, title=" Consultas")
        ctrl.grid(row=1, column=0, sticky="ew", padx=12, pady=8)

        row1 = theme.frame(ctrl, bg=theme.BG_CARD)
        row1.pack(fill="x", padx=10, pady=8)

        theme.label(row1, "Fecha:", bg=theme.BG_CARD).pack(side="left", padx=(0,6))
        self.calendario = DateEntry(
            row1, width=12,
            background=theme.BG_CARD, foreground=theme.TEXT_MAIN,
            headersbackground=theme.ACCENT, headersforeground="white",
            selectbackground=theme.ACCENT,
            date_pattern="yyyy-mm-dd", font=theme.FONT_BODY
        )
        self.calendario.pack(side="left", padx=4)
        theme.btn_primary(row1, "Caja del día",     self.ver_caja_fecha).pack(side="left", padx=4)
        theme.btn_ghost  (row1, "Ganancia neta",    self.ver_ganancia_neta_dia).pack(side="left", padx=4)
        theme.btn_ghost  (row1, "Gráfico métodos",  self.grafico_metodos).pack(side="left", padx=4)
        theme.btn_success(row1, "📥 Exportar Excel", self.exportar_excel_dia).pack(side="left", padx=4)

        theme.separator(ctrl).pack(fill="x", padx=10)

        row2 = theme.frame(ctrl, bg=theme.BG_CARD)
        row2.pack(fill="x", padx=10, pady=8)

        theme.label(row2, "Mes:", bg=theme.BG_CARD).pack(side="left", padx=(0,6))
        self.combo_mes = ttk.Combobox(row2, state="readonly", width=12, font=theme.FONT_BODY)
        self.combo_mes.pack(side="left", padx=4)
        theme.btn_primary(row2, "Ingresos del mes",  self.ver_ganancia_mes).pack(side="left", padx=4)
        theme.btn_ghost  (row2, "Ganancia neta mes", self.ver_ganancia_neta_mes).pack(side="left", padx=4)
        theme.btn_ghost  (row2, "Ver egresos",       self.ver_egresos_mes).pack(side="left", padx=4)

        # ── Resultado ────────────────────────────────
        result_card = theme.card(self.frame, title=" Resultado")
        result_card.grid(row=2, column=0, sticky="nsew", padx=12, pady=8)
        result_card.rowconfigure(0, weight=1)
        result_card.columnconfigure(0, weight=1)

        self.texto = tk.Text(
            result_card,
            font=theme.FONT_MONO,
            bg=theme.BG_CARD, fg=theme.TEXT_MAIN,
            relief="flat", padx=16, pady=12,
            state="disabled"
        )
        self.texto.pack(fill="both", expand=True, padx=8, pady=8)

        self.cargar_meses_disponibles()

    # ─────────────────────────────────────────
    def _mostrar(self, texto):
        self.texto.config(state="normal")
        self.texto.delete("1.0", tk.END)
        self.texto.insert(tk.END, texto)
        self.texto.config(state="disabled")

    def cargar_meses_disponibles(self):
        import sqlite3
        with database.conectar() as conn:
            meses = [r[0] for r in conn.execute(
                "SELECT DISTINCT substr(fecha,1,7) FROM ventas ORDER BY fecha DESC"
            ).fetchall()]
        self.combo_mes["values"] = meses
        if meses:
            self.combo_mes.current(0)

    def ver_caja_fecha(self):
        fecha = self.calendario.get()
        with database.conectar() as conn:
            row = conn.execute(
                "SELECT COUNT(*), COALESCE(SUM(total),0) FROM ventas WHERE date(fecha)=?", (fecha,)
            ).fetchone()
            metodos = conn.execute(
                "SELECT metodo_pago, SUM(total) FROM ventas WHERE date(fecha)=? GROUP BY metodo_pago", (fecha,)
            ).fetchall()
        ingresos, egresos, balance = database.obtener_balance_dia(fecha)

        txt  = f"══════════════ CAJA DEL {fecha} ══════════════\n\n"
        txt += f"  Ventas realizadas:  {row[0]}\n"
        txt += f"  Total cobrado:      ${row[1]:.2f}\n\n"
        txt += "  Por método de pago:\n"
        for m, v in metodos:
            txt += f"    {m:<20} ${v:.2f}\n"
        txt += f"\n  Egresos del día:    ${egresos:.2f}\n"
        txt += f"  Balance neto:       ${balance:.2f}\n"
        self._mostrar(txt)

    def ver_ganancia_neta_dia(self):
        fecha = self.calendario.get()
        with database.conectar() as conn:
            row = conn.execute("""
                SELECT COALESCE(SUM(d.subtotal),0), COALESCE(SUM(p.precio_costo * d.cantidad),0)
                FROM detalle_venta d
                JOIN productos p ON d.producto_id = p.id
                JOIN ventas v ON d.venta_id = v.id
                WHERE date(v.fecha) = ?
            """, (fecha,)).fetchone()
        vendido, costo = row
        _, egresos, _ = database.obtener_balance_dia(fecha)
        ganancia_bruta = vendido - costo
        ganancia_neta = ganancia_bruta - egresos

        txt  = f"══════════════ GANANCIA NETA — {fecha} ══════════════\n\n"
        txt += f"  Total vendido:      ${vendido:.2f}\n"
        txt += f"  Costo mercadería:   ${costo:.2f}\n"
        txt += f"  Ganancia bruta:     ${ganancia_bruta:.2f}\n"
        txt += f"  Egresos del día:    ${egresos:.2f}\n"
        txt += f"  ─────────────────────────────────\n"
        txt += f"  GANANCIA NETA:      ${ganancia_neta:.2f}\n"
        self._mostrar(txt)

    def grafico_metodos(self):
        fecha = self.calendario.get()
        with database.conectar() as conn:
            datos = conn.execute(
                "SELECT metodo_pago, SUM(total) FROM ventas WHERE date(fecha)=? GROUP BY metodo_pago", (fecha,)
            ).fetchall()
        if not datos:
            messagebox.showinfo("Sin datos", f"No hay ventas para el {fecha}")
            return

        colores = [theme.ACCENT, theme.INFO, theme.SUCCESS, theme.ACCENT2, theme.DANGER]
        metodos = [d[0] for d in datos]
        montos  = [d[1] for d in datos]

        fig, ax = plt.subplots(facecolor="#16213e")
        ax.set_facecolor("#0f3460")
        bars = ax.bar(metodos, montos, color=colores[:len(datos)], width=0.5)
        ax.bar_label(bars, fmt="$%.0f", color=theme.TEXT_MAIN, padding=4)
        ax.set_title(f"Métodos de Pago — {fecha}", color=theme.TEXT_MAIN, pad=14)
        ax.tick_params(colors=theme.TEXT_MUTED)
        for spine in ax.spines.values():
            spine.set_edgecolor(theme.BORDER)
        ax.set_ylabel("Monto ($)", color=theme.TEXT_MUTED)
        plt.tight_layout()
        plt.show()

    def exportar_excel_dia(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showerror("Error", "Instalá openpyxl:\n\npip install openpyxl")
            return
        from tkinter import filedialog

        fecha = self.calendario.get()

        with database.conectar() as conn:
            ventas = conn.execute(
                "SELECT id, fecha, total, metodo_pago FROM ventas WHERE date(fecha)=? ORDER BY fecha",
                (fecha,)
            ).fetchall()
            row_tot = conn.execute(
                "SELECT COUNT(*), COALESCE(SUM(total),0) FROM ventas WHERE date(fecha)=?", (fecha,)
            ).fetchone()
            metodos = conn.execute(
                "SELECT metodo_pago, SUM(total) FROM ventas WHERE date(fecha)=? GROUP BY metodo_pago", (fecha,)
            ).fetchall()
            egresos = conn.execute(
                "SELECT tipo, descripcion, monto, metodo FROM egresos WHERE date(fecha)=? ORDER BY fecha",
                (fecha,)
            ).fetchall()
            ingresos, total_egresos, balance = database.obtener_balance_dia(fecha)

        wb = openpyxl.Workbook()

        # ── Estilos ──────────────────────────────────
        def hdr_fill(color):
            return PatternFill("solid", fgColor=color)

        thin = Side(style="thin", color="AAAAAA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style_header(cell, color="1F3864"):
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = hdr_fill(color)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        def style_cell(cell, bold=False, color=None, align="left", number=False):
            cell.font = Font(bold=bold, color=color or "222222")
            cell.alignment = Alignment(horizontal=align)
            cell.border = border
            if number:
                cell.number_format = '"$"#,##0.00'

        # ── Hoja 1: Ventas ───────────────────────────
        ws = wb.active
        ws.title = "Ventas"
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = f"Reporte de Caja — {fecha}"
        title_cell.font = Font(bold=True, size=14, color="1F3864")
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 28

        headers = ["#Venta", "Hora", "Total", "Método de Pago"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            style_header(cell)
        ws.row_dimensions[2].height = 20

        for i, v in enumerate(ventas, 3):
            ws.cell(row=i, column=1, value=v[0]).border = border
            ws.cell(row=i, column=2, value=v[1][11:16]).border = border
            c_total = ws.cell(row=i, column=3, value=v[2])
            style_cell(c_total, align="right", number=True)
            ws.cell(row=i, column=4, value=v[3]).border = border
            if i % 2 == 0:
                for col in range(1, 5):
                    ws.cell(row=i, column=col).fill = hdr_fill("EEF2FF")

        last = len(ventas) + 3
        ws.cell(row=last, column=2, value="TOTAL").font = Font(bold=True)
        c_sum = ws.cell(row=last, column=3, value=row_tot[1])
        style_cell(c_sum, bold=True, color="1F3864", align="right", number=True)
        ws.cell(row=last, column=3).fill = hdr_fill("D9E1F2")

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 20

        # ── Hoja 2: Resumen ──────────────────────────
        ws2 = wb.create_sheet("Resumen")
        ws2.sheet_view.showGridLines = False

        ws2.merge_cells("A1:B1")
        ws2["A1"].value = f"Resumen — {fecha}"
        ws2["A1"].font = Font(bold=True, size=13, color="1F3864")
        ws2["A1"].alignment = Alignment(horizontal="center")
        ws2.row_dimensions[1].height = 26

        resumen = [
            ("Cantidad de ventas", row_tot[0]),
            ("Total cobrado",      row_tot[1]),
        ] + [(f"  · {m}", v) for m, v in metodos] + [
            ("Total egresos",      total_egresos),
            ("Balance neto",       balance),
        ]

        for r, (label, valor) in enumerate(resumen, 2):
            c1 = ws2.cell(row=r, column=1, value=label)
            c2 = ws2.cell(row=r, column=2, value=valor)
            c1.border = border
            c2.border = border
            if label in ("Total cobrado", "Balance neto"):
                style_cell(c1, bold=True)
                style_cell(c2, bold=True, align="right", number=True)
                c1.fill = hdr_fill("D9E1F2")
                c2.fill = hdr_fill("D9E1F2")
            elif label == "Total egresos":
                style_cell(c2, bold=True, color="C00000", align="right", number=True)
                c2.fill = hdr_fill("FCE4D6")
            else:
                c2.alignment = Alignment(horizontal="right")
                if isinstance(valor, float):
                    c2.number_format = '"$"#,##0.00'

        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 16

        # ── Hoja 3: Egresos ──────────────────────────
        ws3 = wb.create_sheet("Egresos")
        ws3.sheet_view.showGridLines = False

        ws3.merge_cells("A1:D1")
        ws3["A1"].value = f"Egresos — {fecha}"
        ws3["A1"].font = Font(bold=True, size=13, color="1F3864")
        ws3["A1"].alignment = Alignment(horizontal="center")

        for col, h in enumerate(["Tipo", "Descripción", "Monto", "Método"], 1):
            style_header(ws3.cell(row=2, column=col))
            ws3.cell(row=2, column=col).value = h

        for i, eg in enumerate(egresos, 3):
            tipo, desc, monto, metodo = eg
            ws3.cell(row=i, column=1, value=tipo).border = border
            ws3.cell(row=i, column=2, value=desc or "").border = border
            c_m = ws3.cell(row=i, column=3, value=monto)
            style_cell(c_m, align="right", number=True)
            ws3.cell(row=i, column=4, value=metodo or "").border = border

        ws3.column_dimensions["A"].width = 20
        ws3.column_dimensions["B"].width = 30
        ws3.column_dimensions["C"].width = 14
        ws3.column_dimensions["D"].width = 16

        # ── Guardar ──────────────────────────────────
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"caja_{fecha}.xlsx"
        )
        if path:
            wb.save(path)
            messagebox.showinfo("✓ Exportado", f"Archivo guardado en:\n{path}")

    def ver_ganancia_mes(self):
        mes = self.combo_mes.get()
        if not mes:
            messagebox.showwarning("Aviso", "No hay meses disponibles")
            return
        with database.conectar() as conn:
            row = conn.execute(
                "SELECT COUNT(*), COALESCE(SUM(total),0) FROM ventas WHERE substr(fecha,1,7)=?", (mes,)
            ).fetchone()
            metodos = conn.execute(
                "SELECT metodo_pago, SUM(total) FROM ventas WHERE substr(fecha,1,7)=? GROUP BY metodo_pago", (mes,)
            ).fetchall()
        txt  = f"══════════════ INGRESOS — {mes} ══════════════\n\n"
        txt += f"  Ventas totales:     {row[0]}\n"
        txt += f"  Ingreso total:      ${row[1]:.2f}\n\n"
        txt += "  Por método:\n"
        for m, v in metodos:
            txt += f"    {m:<20} ${v:.2f}\n"
        self._mostrar(txt)

    def ver_ganancia_neta_mes(self):
        mes = self.combo_mes.get()
        if not mes:
            messagebox.showwarning("Aviso", "No hay meses disponibles")
            return
        with database.conectar() as conn:
            row = conn.execute("""
                SELECT COALESCE(SUM(d.subtotal),0), COALESCE(SUM(p.precio_costo * d.cantidad),0)
                FROM detalle_venta d
                JOIN productos p ON d.producto_id = p.id
                JOIN ventas v ON d.venta_id = v.id
                WHERE substr(v.fecha,1,7) = ?
            """, (mes,)).fetchone()
            eg = conn.execute(
                "SELECT COALESCE(SUM(monto),0) FROM egresos WHERE substr(fecha,1,7)=?", (mes,)
            ).fetchone()[0]
        vendido, costo = row
        ganancia = vendido - costo - eg

        txt  = f"══════════════ GANANCIA NETA — {mes} ══════════════\n\n"
        txt += f"  Total vendido:      ${vendido:.2f}\n"
        txt += f"  Costo mercadería:   ${costo:.2f}\n"
        txt += f"  Egresos del mes:    ${eg:.2f}\n"
        txt += f"  ─────────────────────────────────\n"
        txt += f"  GANANCIA NETA:      ${ganancia:.2f}\n"
        self._mostrar(txt)

    def ver_egresos_mes(self):
        mes = self.combo_mes.get()
        if not mes:
            return
        egresos = database.obtener_egresos_por_mes(mes)
        txt = f"══════════════ EGRESOS — {mes} ══════════════\n\n"
        if not egresos:
            txt += "  Sin egresos registrados.\n"
        else:
            total = 0
            for eg in egresos:
                eid, tipo, desc, monto, metodo, fecha = eg
                txt += f"  {fecha[:10]}  [{tipo}]  {desc or ''}   ${monto:.2f}  ({metodo})\n"
                total += monto
            txt += f"\n  ─────────────────────────────────\n"
            txt += f"  TOTAL EGRESOS: ${total:.2f}\n"
        self._mostrar(txt)
